import argparse
import csv
import re
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET
from urllib.parse import parse_qs, urlparse


def normalize_asset_id(value):
    raw = "" if value is None else str(value).strip()
    if raw == "":
        return raw, "", "empty", "empty"
    if raw.startswith("rbxassetid://") and raw[len("rbxassetid://") :].isdigit():
        return raw, raw, "ok", "prefixed"
    if raw.isdigit():
        return raw, f"rbxassetid://{raw}", "ok", "numeric"
    parsed = urlparse(raw)
    query_id = parse_qs(parsed.query).get("id", [""])[0]
    if query_id.isdigit():
        return raw, f"rbxassetid://{query_id}", "ok", "url"
    return raw, "", "bad", "unknown"


def read_rows(table_path):
    suffix = table_path.suffix.lower()
    if suffix == ".csv":
        with table_path.open("r", encoding="utf-8-sig", newline="") as f:
            return list(csv.DictReader(f))

    if suffix == ".xlsx":
        try:
            import openpyxl
        except ImportError:
            return read_xlsx_with_stdlib(table_path)

        wb = openpyxl.load_workbook(table_path, data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        raw_rows = list(ws.iter_rows(values_only=True))
        if not raw_rows:
            return []
        headers = ["" if v is None else str(v).strip() for v in raw_rows[0]]
        rows = []
        for raw in raw_rows[1:]:
            row = {}
            for index, header in enumerate(headers):
                if header:
                    row[header] = "" if index >= len(raw) or raw[index] is None else str(raw[index])
            rows.append(row)
        return rows

    raise SystemExit(f"Unsupported table file: {table_path}")


def read_xlsx_with_stdlib(table_path):
    ns = {
        "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
        "rel": "http://schemas.openxmlformats.org/package/2006/relationships",
    }

    with zipfile.ZipFile(table_path) as zf:
        shared_strings = []
        if "xl/sharedStrings.xml" in zf.namelist():
            root = ET.fromstring(zf.read("xl/sharedStrings.xml"))
            for item in root.findall("main:si", ns):
                parts = [node.text or "" for node in item.findall(".//main:t", ns)]
                shared_strings.append("".join(parts))

        sheet_path = "xl/worksheets/sheet1.xml"
        if "xl/workbook.xml" in zf.namelist() and "xl/_rels/workbook.xml.rels" in zf.namelist():
            book_root = ET.fromstring(zf.read("xl/workbook.xml"))
            first_sheet = book_root.find("main:sheets/main:sheet", ns)
            rel_id = first_sheet.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id") if first_sheet is not None else ""
            rel_root = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
            for rel in rel_root.findall("rel:Relationship", ns):
                if rel.attrib.get("Id") == rel_id:
                    target = rel.attrib.get("Target", "")
                    clean_target = target.lstrip("/")
                    sheet_path = clean_target if clean_target.startswith("xl/") else "xl/" + clean_target
                    break

        if sheet_path not in zf.namelist():
            return []

        sheet_root = ET.fromstring(zf.read(sheet_path))
        raw_rows = []
        for row in sheet_root.findall(".//main:sheetData/main:row", ns):
            values = []
            for cell in row.findall("main:c", ns):
                ref = cell.attrib.get("r", "")
                match = re.match(r"([A-Z]+)", ref)
                if match:
                    col = 0
                    for ch in match.group(1):
                        col = col * 26 + ord(ch) - ord("A") + 1
                    while len(values) < col - 1:
                        values.append("")

                value = ""
                cell_type = cell.attrib.get("t", "")
                if cell_type == "inlineStr":
                    parts = [node.text or "" for node in cell.findall(".//main:t", ns)]
                    value = "".join(parts)
                else:
                    node = cell.find("main:v", ns)
                    value = node.text if node is not None and node.text is not None else ""
                    if cell_type == "s" and value.isdigit():
                        index = int(value)
                        value = shared_strings[index] if index < len(shared_strings) else ""
                values.append(value)
            raw_rows.append(values)

    if not raw_rows:
        return []
    headers = [str(v).strip() for v in raw_rows[0]]
    rows = []
    for raw in raw_rows[1:]:
        row = {}
        for index, header in enumerate(headers):
            if header:
                row[header] = "" if index >= len(raw) else str(raw[index])
        rows.append(row)
    return rows


def md_cell(value):
    return str(value or "").replace("|", "\\|")


def main():
    parser = argparse.ArgumentParser(description="Check Roblox UI asset map tables.")
    parser.add_argument("--table", required=True, help="CSV or XLSX asset map path")
    parser.add_argument("--asset-root", default="", help="Root folder for local PNG files")
    parser.add_argument("--out", required=True, help="Markdown report output path")
    parser.add_argument("--focus", default="", help="Comma-separated rel_path list")
    args = parser.parse_args()

    table_path = Path(args.table).resolve()
    asset_root = Path(args.asset_root).resolve() if args.asset_root else None
    out_path = Path(args.out).resolve()
    focus = {p.strip().replace("\\", "/").lower() for p in args.focus.split(",") if p.strip()}

    rows = read_rows(table_path)
    checked = []
    for row in rows:
        rel_path = str(row.get("rel_path", "")).strip()
        if not rel_path:
            continue
        rel_key = rel_path.replace("\\", "/").lower()
        if focus and rel_key not in focus:
            continue

        usage_hint = str(row.get("usage_hint", ""))
        is_reference = "reference_only_layout_screenshot" in usage_hint or "no_asset_id_needed" in usage_hint
        raw, normalized, state, fmt = normalize_asset_id(row.get("asset_id", ""))
        local_path = asset_root / rel_path if asset_root else None
        local_exists = local_path.exists() if local_path else False

        status = "ready"
        if is_reference:
            status = "reference_only"
        elif state == "bad":
            status = "bad_asset_id"
        elif state == "empty":
            status = "missing_asset_id"
        elif asset_root and not local_exists:
            status = "local_missing"

        checked.append(
            {
                "status": status,
                "rel_path": rel_path,
                "asset_id_raw": raw,
                "asset_id": normalized,
                "format": fmt,
                "local_exists": local_exists,
            }
        )

    counts = {}
    for row in checked:
        counts[row["status"]] = counts.get(row["status"], 0) + 1

    lines = [
        "# UI AssetId Map Check",
        "",
        f"TablePath: `{table_path}`",
        f"AssetRoot: `{asset_root}`" if asset_root else "AssetRoot: ``",
        "",
        "## Summary",
        "",
    ]
    for key in sorted(counts):
        lines.append(f"- {key}: {counts[key]}")

    needs = [row for row in checked if row["status"] in {"missing_asset_id", "bad_asset_id", "local_missing"}]
    lines += ["", "## Rows Needing Confirmation", ""]
    if needs:
        lines += ["| status | rel_path | asset_id_raw | normalized_asset_id | local_exists |", "|---|---|---|---|---|"]
        for row in needs:
            lines.append(
                f"| {md_cell(row['status'])} | {md_cell(row['rel_path'])} | {md_cell(row['asset_id_raw'])} | {md_cell(row['asset_id'])} | {row['local_exists']} |"
            )
    else:
        lines.append("None.")

    lines += [
        "",
        "## Checked Rows",
        "",
        "| status | rel_path | asset_id_raw | normalized_asset_id | format | local_exists |",
        "|---|---|---|---|---|---|",
    ]
    for row in checked:
        lines.append(
            f"| {md_cell(row['status'])} | {md_cell(row['rel_path'])} | {md_cell(row['asset_id_raw'])} | {md_cell(row['asset_id'])} | {md_cell(row['format'])} | {row['local_exists']} |"
        )

    lines += [
        "",
        "## Codex Rule",
        "",
        "- `asset_id` can be `123456`, `rbxassetid://123456`, or a Roblox URL containing `id=123456`.",
        "- If XLSX and CSV disagree, use the user-specified source table first, then sync or regenerate the secondary file.",
    ]

    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(f"Wrote asset map check: {out_path}")
    for key in sorted(counts):
        print(f"{key}: {counts[key]}")


if __name__ == "__main__":
    main()
